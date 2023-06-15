import requests
from bs4 import BeautifulSoup
import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import openpyxl.utils
import time
from openpyxl.styles import PatternFill

# 글의 목록을 조회, 세부정보로 넘어가는 링크 수집
def get_link(start_date):
    count = 1
    href_list = []
    start_date = f"{start_date.year}%2F{start_date.month}%2F{start_date.day}"
    last_date = datetime.datetime.now()
    last_date = f"{last_date.year}%2F{last_date.month}%2F{last_date.day}"
    while(True):
        reqUrl = f"https://www.g2b.go.kr:8101/ep/tbid/tbidList.do?searchType=1&bidSearchType=1&searchDtType=1&fromBidDt={start_date}&toBidDt={last_date}&setMonth1=1&radOrgan=1&industryCd=1426&recordCountPerPage=100&currentPageNo={count}"
        headersList = {
        "Accept": "*/*",
        "User-Agent": "Thunder Client (https://www.thunderclient.com)" 
        }

        payload = ""

        response = requests.request("GET", reqUrl, data=payload,  headers=headersList)
        soup = BeautifulSoup(response.text, 'html.parser')
        link_list = soup.find_all('td',class_='tl')
        for link in link_list:
            a_tag = link.find('a')
            if a_tag:
                href = a_tag.get('href')
                href_list.append(href)
        count += 1
        if len(link_list) == 0:
            break
        else:
            link_list = []
    return href_list

# 세부정보의 내용 수집
def get_impormation(href):
    response = requests.get(href)
    soup = BeautifulSoup(response.text, 'html.parser')
    result = None
    table = soup.find('table',summary='공고일반 정보')
    if table is None:
        table = soup.find_all('table', class_='table_info')
        for tab in table:
            result = tab.find('span', text= '공고일반 정보')
            if result is not None:
                table = tab
                break
            else:
                table = None
    tr_list = table.find_all('tr')

    number = None
    name = None
    institution = None
    agency = None

    general_info = {}
    for tr in tr_list:
        th_list = tr.find_all('th')
        for th in th_list:
            if "입찰공고번호" in th.text:
                number = th.find_next_sibling('td').text.strip()
            elif "공고명" in th.text:
                name = th.find_next_sibling('td').text.strip().replace('\r\n','').replace("  ","")
            elif "공고기관" in th.text:
                institution = th.find_next_sibling('td').text.strip()
            elif "수요기관" in th.text:
                agency = th.find_next_sibling('td').text.strip()

    general_info["입찰공고번호"] = number
    general_info["공고명"] = name
    general_info["공고기관"] = institution
    general_info["수요기관"] = agency


    bid_start = ''
    bid_end = ''
    result = None
    table = soup.find('table',summary='입찰집행 및 진행 정보')
    if table is None:
        table = soup.find_all('table', class_='table_info')
        for tab in table:
            result = tab.find('span', text= '입찰집행 및 진행 정보')
            if result is not None:
                table = tab
                break
            else:
                table = None
    
    if table is None:
        table = soup.find('table',summary = '기본제안서 제출 및 공동수급협정 정보')
            
    tr_list = table.find_all('tr')
    for tr in tr_list:
        th_list = tr.find_all('th')
        for th in th_list:
            if "입찰개시일시" in th.text:
                bid_start = th.find_next_sibling('td').text.strip()
            if "입찰마감일시" in th.text:
                bid_end = th.find_next_sibling('td').text.strip()
    bid_info = {"입찰개시일시":bid_start, "입찰마감일시":bid_end}

    result = None
    table = soup.find('table',summary='예정가격 결정 및 입찰금액 정보')
    if table is None:
        table = soup.find_all('table', class_='table_info')
        for tab in table:
            result = tab.find('span', text= '입찰금액')
            if result is not None:
                table = tab
                break
            else:
                table = None
    
    if table is None:
        table = soup.find_all('table', class_='table_info')
        for tab in table:
            result = tab.find('span', text= '예정가격 결정 및 입찰금액 정보')
            if result is not None:
                table = tab
                break
            else:
                table = None
    

    tr_list = table.find_all('tr')
    
    money_info = {}
    money = None
    calcul_money = None
    for tr in tr_list:
        th_list = tr.find_all('th')
        for th in th_list:
            if "사업금액" in th.text:
                money = th.find_next_sibling('td').text.strip()
            if "추정가격" in th.text:
                calcul_money = th.find_next_sibling('td').text.split('(')[0].strip()

    money_info["사업금액"] = money
    money_info["추정가격"] = calcul_money

    result = None
    table = soup.find('table',summary='첨부 파일 정보')
    if table is None or table == []:
        table = soup.find_all('table', class_='table_list_attchFileTbl')
        for tab in table:
            result = tab.find('span', text= '파일')
            if result is not None:
                table = tab
                break
            else:
                table = None
    
    if table is None or table == []:
        infomation = {'공고일반정보':general_info, '입찰집행 및 진행 정보':bid_info, '예정가격 결정 및 입찰금액 정보':money_info}
        return infomation
    
    a_list = table.find_all('a')
    down_info = []
    for a in a_list:
        file_info = {}
        if result is not None:
            file_info['파일명'] = a.get('href').split('d(\'')[1].split('\',')[1].split('\');')[0]
            file_info['링크'] = a.get('href').split('d(\'')[1].split('\',')[0]
        else:
            file_info['파일명'] = a.get('href').split('d(\'')[1].split('\', \'')[1].split('\');')[0]
            file_info['링크'] = a.get('href').split('d(\'')[1].split('\', ')[0]
        down_info.append(file_info)
    infomation = {'공고일반정보':general_info, '입찰집행 및 진행 정보':bid_info, '예정가격 결정 및 입찰금액 정보':money_info, '첨부 파일 정보':down_info}
    
    return infomation

# 첨부파일 다운로드
def get_file(down_info, number):
    current_directory = os.getcwd()
    first_directory = os.path.join(current_directory, 'Downloads')
    
    # 디렉토리가 없는 경우 디렉토리를 생성
    if not os.path.exists(first_directory):
        os.makedirs(first_directory)

    end_directort = os.path.join(first_directory, number)
    if not os.path.exists(end_directort):
        os.makedirs(end_directort)
    
    for info in down_info:
        filename = info['파일명']
        link = info['링크']
        filename = os.path.join(end_directort, filename)
        link =f'https://www.g2b.go.kr:8081/ep/co/fileDownload.do?fileTask=NOTIFY&fileSeq={link}'
        with open(filename, "wb") as file:
            response = requests.get(link)
            file.write(response.content)
        return end_directort

def writer(href, header, data, filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(header)

    for d in data:
        row = []
        for item in d:
            if item == d[1]:
                row.append(f'=HYPERLINK("{href}", "{item}")')
            elif 'file://' in str(item):
                row.append(f'=HYPERLINK("{item}", "첨부파일(클릭 시 폴더로 이동합니다.)")')
            else:
                row.append(item)
        ws.append(row)

        column_widths = [18, 60, 20, 20, 18, 18, 14, 14, 33]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width

        for cell in ws[ws.max_row]: 
            cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
    wb.save(filename)

def read_first_column(filename):
    wb = load_workbook(filename)
    ws = wb.active

    column_values = []
    for cell in ws['A']:
        column_values.append(cell.value)

    return column_values

# 실행  
if __name__ == "__main__":
    print('나라장터 수집기는 나라장터의 소프트웨어사업자 업종의 용역 정보만을 수집합니다.')
    print('첫 수집 실행시 엑소스피어에 의해 차단되며 우측 하단에 발생하는 엑소스피어 알림에서 허용 후 정상 작동합니다.')
    print('날짜 입력 전 엑셀 파일을 종료시켜주십시오.')
    print('수집 시작날짜 입력(ex) 20230601):')
    date = input()
    year = int(date[0:4])
    month = int(date[4:6])
    day = int(date[6:8])
    try:
        start_date = datetime.date(year,month,day)
    except ValueError as e:
        if 'day is out of range for month' in str(e):
            print('잘못된 날짜를 입력하셨습니다. 3초후에 종료됩니다.')
            time.sleep(3)
            quit()
    print('수집 시작')
    filename = "나라장터 용역 리스트.xlsx"
    header = ["공고번호", "공고명(클릭 시 이동)", "공고기관", "수요기관", "입찰개시일시", "입찰마감일시","사업금액", "추정금액", "첨부파일(클릭 시 이동)"]
    href_list = get_link(start_date)
    count_all = len(href_list)
    complete_count = 0
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        
    fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
    for cell in ws[1]:
        cell.fill = fill
        
    wb.save(filename)

    for href in href_list:
        already = read_first_column(filename)
        infomation = get_impormation(href)
        
        if infomation['공고일반정보']['입찰공고번호'] in already:
            complete_count += 1
            print(f"수집중({complete_count}/{count_all})")
            continue
        
        if infomation.get('첨부 파일 정보') is not None and infomation.get('첨부 파일 정보') != []:
            down_info = infomation['첨부 파일 정보']
            down_path = get_file(down_info, infomation['공고일반정보']['입찰공고번호'])
            data = [[infomation['공고일반정보']['입찰공고번호'],infomation['공고일반정보']['공고명'],
                     infomation['공고일반정보']['공고기관'],infomation['공고일반정보']['수요기관'],
                     infomation['입찰집행 및 진행 정보']['입찰개시일시'],infomation['입찰집행 및 진행 정보']['입찰마감일시'],
                     infomation['예정가격 결정 및 입찰금액 정보']['사업금액'],infomation['예정가격 결정 및 입찰금액 정보']['추정가격'],
                     'file://'+ down_path
                     ]]
            writer(href,header, data, filename)
        else:
            data = [[infomation['공고일반정보']['입찰공고번호'],infomation['공고일반정보']['공고명'],
                     infomation['공고일반정보']['공고기관'],infomation['공고일반정보']['수요기관'],
                     infomation['입찰집행 및 진행 정보']['입찰개시일시'],infomation['입찰집행 및 진행 정보']['입찰마감일시'],
                     infomation['예정가격 결정 및 입찰금액 정보']['사업금액'],infomation['예정가격 결정 및 입찰금액 정보']['추정가격'],
                     "첨부파일 없음"]]
            writer(href,header, data, filename)
        complete_count += 1
        print(f"수집중({complete_count}/{count_all})")
        
    print('수집이 완료되었습니다. 3초뒤에 종료됩니다.')
    time.sleep(3)
    quit()